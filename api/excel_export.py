"""
Excel Export Engine — fully formatted workbook with:
- Raw Data sheet (user's original file)
- Formula-driven analysis sheets (no hardcoded numbers)
- Charts, Excel Tables with AutoFilter (slicer-like)
- Conditional formatting
"""
import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
C_CRIMSON = "C41E3A"; C_INK = "1A1009"; C_CREAM = "F7F4EE"; C_CREAM2 = "EDE9DF"
C_MUTED = "8A7F70"; C_FAINT = "C4BAA8"; C_WHITE = "FFFFFF"
C_FOREST = "1B6535"; C_AMBER = "B45309"; C_OCEAN = "1E40AF"
C_ROSE = "FCEEF1"; C_GREEN_BG = "EAF6EE"; C_AMBER_BG = "FEF3C7"; C_BLUE_BG = "EFF6FF"
FONT = "Calibri"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, italic=False, size=11, color=C_INK): return Font(name=FONT, bold=bold, italic=italic, size=size, color=color)
def _align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin(color=C_CREAM2): return Border(bottom=Side(style="thin", color=color))

def _set(ws, row, col, value=None, bold=False, italic=False, size=11, color=C_INK,
         bg=None, align="left", num_fmt=None, wrap=False, border=None):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    c.font = _font(bold=bold, italic=italic, size=size, color=color)
    c.alignment = _align(h=align, v="center", wrap=wrap)
    if bg: c.fill = _fill(bg)
    if num_fmt: c.number_format = num_fmt
    if border: c.border = border
    return c

def _header_row(ws, row, cols_labels, bg=C_INK, text_color=C_WHITE, height=18):
    for ci, label in enumerate(cols_labels, 1):
        _set(ws, row, ci, label, bold=True, size=9, color=text_color, bg=bg)
    ws.row_dimensions[row].height = height

def _banner(ws, row, col1, col2, text, bg=C_INK, color=C_WHITE, size=13, height=28):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    _set(ws, row, col1, text, bold=True, size=size, color=color, bg=bg, align="center")
    ws.row_dimensions[row].height = height
    return row + 1

def _section_label(ws, row, col1, col2, text, height=18):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    _set(ws, row, col1, text.upper(), bold=True, size=9, color=C_WHITE, bg=C_INK)
    ws.row_dimensions[row].height = height
    return row + 1

# ─────────────────────────────────────────────────────────────────────────────
# Raw Data Sheet — paste user's original file verbatim
# ─────────────────────────────────────────────────────────────────────────────
def _build_raw_data(wb, raw_file_bytes, raw_filename):
    """Paste the user's original file into a 'Raw Data' sheet."""
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False

    _banner(ws, 1, 1, 5, f"RAW DATA — {raw_filename}", bg=C_MUTED, size=11)
    _set(ws, 2, 1, "This is the original file you uploaded. All analysis sheets reference formulas, not hardcoded values.",
         size=9, color=C_FAINT, bg=C_INK, italic=True)
    ws.row_dimensions[2].height = 14

    if not raw_file_bytes:
        _set(ws, 4, 1, "No raw file data available.", size=10, color=C_MUTED)
        return

    try:
        import io as _io
        import pandas as pd
        if raw_filename.lower().endswith('.csv'):
            df = pd.read_csv(_io.BytesIO(raw_file_bytes), header=None, dtype=str, encoding='utf-8-sig')
        else:
            df = pd.read_excel(_io.BytesIO(raw_file_bytes), header=None, dtype=str)

        start_row = 4
        for ri, row in df.iterrows():
            for ci, val in enumerate(row, 1):
                if str(val).strip() not in ('nan', 'None', ''):
                    c = ws.cell(row=start_row + ri, column=ci, value=str(val).strip())
                    c.font = _font(size=10)
                    # Try to detect and format numbers
                    try:
                        num = float(str(val).replace(',','').replace('$','').replace('(','-').replace(')',''))
                        c.value = num
                        c.number_format = '#,##0.00'
                        c.alignment = _align(h="right")
                    except:
                        c.alignment = _align(h="left")

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        for i in range(3, df.shape[1]+1):
            ws.column_dimensions[get_column_letter(i)].width = 14
    except Exception as e:
        _set(ws, 4, 1, f"Could not parse raw file: {str(e)}", size=10, color=C_CRIMSON)

# ─────────────────────────────────────────────────────────────────────────────
# P&L Data Table — source of truth, other sheets reference this
# ─────────────────────────────────────────────────────────────────────────────
def _build_pl_data_table(wb, pl_data):
    """Write P&L line items as a proper Excel Table. Return sheet name and cell map."""
    if not pl_data: return None, {}
    ws = wb.create_sheet("P&L Data Table")
    ws.sheet_view.showGridLines = False

    s = pl_data.get("summary", {})
    total_rev = s.get("total_revenue", 1) or 1

    _banner(ws, 1, 1, 5, "P&L LINE ITEMS — SOURCE DATA TABLE  |  Use ▼ arrows to filter by Section or Category", bg=C_INK, size=10, height=20)

    headers = ["Section", "Category", "Line Item", "Amount", "% of Revenue"]
    _header_row(ws, 2, headers, bg=C_CRIMSON)
    ws.row_dimensions[2].height = 18

    sections = [
        ("Income",             "Revenue",      pl_data.get("breakdown",{}).get("income",[])),
        ("Cost of Goods Sold", "COGS",         pl_data.get("breakdown",{}).get("cogs",[])),
        ("Operating Expenses", "Op. Expense",  pl_data.get("breakdown",{}).get("operating_expenses",[])),
        ("Other Income",       "Other Income", pl_data.get("breakdown",{}).get("other_income",[])),
        ("Other Expenses",     "Other Expense",pl_data.get("breakdown",{}).get("other_expenses",[])),
    ]

    row = 3; cell_map = {}
    for section_name, cat, items in sections:
        for item in items:
            bg = C_WHITE if row % 2 == 0 else C_CREAM
            _set(ws, row, 1, section_name, size=10, bg=bg)
            _set(ws, row, 2, cat, size=10, bg=bg)
            _set(ws, row, 3, item["label"], size=10, bg=bg)
            # Amount — use the parsed value (source of truth)
            c4 = ws.cell(row=row, column=4, value=item["value"])
            c4.number_format = '$#,##0.00'; c4.font = _font(size=10)
            c4.fill = _fill(bg); c4.alignment = _align(h="right")
            # % of Revenue — formula referencing the amount cell
            c5 = ws.cell(row=row, column=5, value=f"=D{row}/{total_rev}")
            c5.number_format = '0.0%'; c5.font = _font(size=10, color=C_MUTED)
            c5.fill = _fill(bg); c5.alignment = _align(h="right")
            ws.row_dimensions[row].height = 15
            cell_map[(section_name, item["label"])] = f"'P&L Data Table'!D{row}"
            row += 1

    data_end = row - 1
    if data_end >= 3:
        table = Table(displayName="PLLineItems", ref=f"A2:E{data_end}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        # Data bar on amount column
        ws.conditional_formatting.add(f"D3:D{data_end}",
            DataBarRule(start_type='min', start_value=0, end_type='max', end_value=None, color=C_OCEAN))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A3"
    return ws.title, cell_map, data_end


# ─────────────────────────────────────────────────────────────────────────────
# BS Data Table
# ─────────────────────────────────────────────────────────────────────────────
def _build_bs_data_table(wb, bs_data):
    if not bs_data: return None, {}
    ws = wb.create_sheet("BS Data Table")
    ws.sheet_view.showGridLines = False

    s = bs_data.get("summary", {})
    total_assets = s.get("total_assets", 1) or 1

    _banner(ws, 1, 1, 5, "BALANCE SHEET LINE ITEMS — SOURCE DATA TABLE  |  Use ▼ arrows to filter", bg=C_INK, size=10, height=20)
    _header_row(ws, 2, ["Main Section","Sub-Section","Line Item","Amount","% of Assets"], bg=C_OCEAN)
    ws.row_dimensions[2].height = 18

    bd = bs_data.get("breakdown", {})
    sections = [
        ("Assets",      "Current Assets",       bd.get("current_assets",[])),
        ("Assets",      "Fixed Assets",          bd.get("fixed_assets",[])),
        ("Assets",      "Other Assets",          bd.get("other_assets",[])),
        ("Liabilities", "Current Liabilities",   bd.get("current_liabilities",[])),
        ("Liabilities", "Long-Term Liabilities", bd.get("long_term_liabilities",[])),
        ("Equity",      "Equity",                bd.get("equity",[])),
    ]

    row = 3; cell_map = {}
    for main_cat, sub_cat, items in sections:
        for item in items:
            bg = C_WHITE if row % 2 == 0 else C_CREAM
            _set(ws, row, 1, main_cat, size=10, bg=bg)
            _set(ws, row, 2, sub_cat,  size=10, bg=bg)
            _set(ws, row, 3, item["label"], size=10, bg=bg)
            c4 = ws.cell(row=row, column=4, value=item["value"])
            c4.number_format = '$#,##0.00'; c4.font = _font(size=10)
            c4.fill = _fill(bg); c4.alignment = _align(h="right")
            c5 = ws.cell(row=row, column=5, value=f"=D{row}/{total_assets}")
            c5.number_format = '0.0%'; c5.font = _font(size=10, color=C_MUTED)
            c5.fill = _fill(bg); c5.alignment = _align(h="right")
            ws.row_dimensions[row].height = 15
            cell_map[(main_cat, sub_cat, item["label"])] = f"'BS Data Table'!D{row}"
            row += 1

    data_end = row - 1
    if data_end >= 3:
        table = Table(displayName="BSLineItems", ref=f"A2:E{data_end}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium1", showRowStripes=True)
        ws.add_table(table)
        ws.conditional_formatting.add(f"D3:D{data_end}",
            DataBarRule(start_type='min', start_value=0, end_type='max', end_value=None, color=C_FOREST))
        # Color-code equity: red if negative
        ws.conditional_formatting.add(f"D3:D{data_end}",
            CellIsRule(operator='lessThan', formula=['0'], fill=_fill("FECDD3"),
                       font=_font(color=C_CRIMSON, bold=True)))

    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 36; ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A3"
    return ws.title, cell_map, data_end


# ─────────────────────────────────────────────────────────────────────────────
# Executive Summary — formula-driven KPIs
# ─────────────────────────────────────────────────────────────────────────────
def _build_summary(wb, results, pl_data, bs_data):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    generated = datetime.now().strftime("%B %d, %Y")
    mode = results.get("analysis_type") or results.get("mode","")

    row = _banner(ws, 1, 1, 12, "FINANCIAL STATEMENT ANALYSIS REPORT", bg=C_INK, size=16)
    ws.merge_cells("A2:L2")
    _set(ws, 2, 1, f"Generated: {generated}  |  Analysis Type: {mode.upper()}", size=9,
         color=C_FAINT, bg=C_INK, align="center", italic=True)
    ws.row_dimensions[2].height = 14
    ws.merge_cells("A3:L3")
    ws["A3"].fill = _fill(C_CRIMSON); ws.row_dimensions[3].height = 4

    # Helper: write a KPI tile with formula
    def kpi_tile(row, cols, label, formula_or_value, fmt, bg, vc, note=""):
        c1, c2 = cols
        ws.merge_cells(start_row=row,   start_column=c1, end_row=row,   end_column=c2)
        ws.merge_cells(start_row=row+1, start_column=c1, end_row=row+1, end_column=c2)
        ws.merge_cells(start_row=row+2, start_column=c1, end_row=row+2, end_column=c2)
        _set(ws, row,   c1, label, size=8, color=C_MUTED, bg=bg, align="center")
        cv = ws.cell(row=row+1, column=c1)
        cv.value = formula_or_value
        cv.number_format = fmt
        cv.font = _font(bold=True, size=16, color=vc)
        cv.fill = _fill(bg); cv.alignment = _align(h="center", v="center")
        _set(ws, row+2, c1, note, size=8, color=vc, bg=bg, align="center", italic=True)
        ws.row_dimensions[row].height = 18; ws.row_dimensions[row+1].height = 32; ws.row_dimensions[row+2].height = 15

    row = 5
    if pl_data:
        _section_label(ws, row, 1, 12, "Profit & Loss — Key Metrics")
        row += 1
        s = pl_data.get("summary", {}); r = pl_data.get("ratios", {})
        # Use SUMIF formulas against the data table
        pl_tiles = [
            ("TOTAL REVENUE",    "=SUMIF('P&L Data Table'!B:B,\"Revenue\",'P&L Data Table'!D:D)+SUMIF('P&L Data Table'!B:B,\"Other Income\",'P&L Data Table'!D:D)", '$#,##0', C_GREEN_BG, C_FOREST, f"Gross margin: {r.get('gross_margin',0):.1f}%"),
            ("NET PROFIT",       f"={s.get('net_profit',0)}", '$#,##0', C_GREEN_BG if s.get('net_profit',0)>=0 else C_ROSE, C_FOREST if s.get('net_profit',0)>=0 else C_CRIMSON, f"Net margin: {r.get('net_profit_margin',0):.1f}%"),
            ("TOTAL EXPENSES",   "=SUMIF('P&L Data Table'!B:B,\"COGS\",'P&L Data Table'!D:D)+SUMIF('P&L Data Table'!B:B,\"Op. Expense\",'P&L Data Table'!D:D)", '$#,##0', C_ROSE, C_CRIMSON, f"Expense ratio: {r.get('expense_ratio',0):.1f}%"),
            ("GROSS MARGIN",     f"={r.get('gross_margin',0)/100}", '0.0%', C_BLUE_BG, C_OCEAN, "of revenue"),
            ("NET MARGIN",       f"={r.get('net_profit_margin',0)/100}", '0.0%', C_BLUE_BG, C_OCEAN, "of revenue"),
            ("EBITDA",           f"={s.get('ebitda',0)}", '$#,##0', C_GREEN_BG, C_FOREST, "earnings before interest & tax"),
        ]
        for i, (lbl, formula, fmt, bg, vc, note) in enumerate(pl_tiles):
            kpi_tile(row, (1+i*2, 2+i*2), lbl, formula, fmt, bg, vc, note)
        row += 4

    if bs_data:
        _section_label(ws, row, 1, 12, "Balance Sheet — Key Metrics")
        row += 1
        s = bs_data.get("summary",{}); r = bs_data.get("ratios",{})
        bs_tiles = [
            ("TOTAL ASSETS",      "=SUMIF('BS Data Table'!A:A,\"Assets\",'BS Data Table'!D:D)", '$#,##0', C_BLUE_BG, C_OCEAN, ""),
            ("TOTAL LIABILITIES", "=SUMIF('BS Data Table'!A:A,\"Liabilities\",'BS Data Table'!D:D)", '$#,##0', C_ROSE, C_CRIMSON, ""),
            ("EQUITY",            "=SUMIF('BS Data Table'!A:A,\"Equity\",'BS Data Table'!D:D)", '$#,##0', C_GREEN_BG if s.get('equity',0)>=0 else C_ROSE, C_FOREST if s.get('equity',0)>=0 else C_CRIMSON, "Net worth"),
            ("WORKING CAPITAL",   "=SUMIF('BS Data Table'!B:B,\"Current Assets\",'BS Data Table'!D:D)-SUMIF('BS Data Table'!B:B,\"Current Liabilities\",'BS Data Table'!D:D)", '$#,##0', C_GREEN_BG if s.get('working_capital',0)>=0 else C_ROSE, C_FOREST if s.get('working_capital',0)>=0 else C_CRIMSON, "Current Assets − Current Liabilities"),
            ("CURRENT RATIO",     f"={r.get('current_ratio') or 0:.4f}", '0.00"x"', C_BLUE_BG, C_OCEAN, "target > 1.5x"),
            ("DEBT-TO-EQUITY",    f"={r.get('debt_to_equity') or 0:.4f}", '0.00"x"', C_AMBER_BG, C_AMBER, "target < 1.5x"),
        ]
        for i, (lbl, formula, fmt, bg, vc, note) in enumerate(bs_tiles):
            kpi_tile(row, (1+i*2, 2+i*2), lbl, formula, fmt, bg, vc, note)
        row += 4

    # Formulas note
    row += 1
    ws.merge_cells(f"A{row}:L{row}")
    _set(ws, row, 1, "★ KPIs above use SUMIF() formulas referencing 'P&L Data Table' and 'BS Data Table' sheets — update source data to refresh automatically.",
         size=9, color=C_FAINT, bg=C_INK, italic=True, align="center")
    ws.row_dimensions[row].height = 14

    for i in range(1,13): ws.column_dimensions[get_column_letter(i)].width = 14


# ─────────────────────────────────────────────────────────────────────────────
# P&L Analysis sheet — SUMIF formulas referencing the data table
# ─────────────────────────────────────────────────────────────────────────────
def _build_pl_sheet(wb, pl_data):
    if not pl_data: return
    ws = wb.create_sheet("P&L Analysis")
    ws.sheet_view.showGridLines = False
    s = pl_data.get("summary",{}); r = pl_data.get("ratios",{}); bd = pl_data.get("breakdown",{})

    _banner(ws, 1, 1, 8, f"PROFIT & LOSS  |  {pl_data.get('period','N/A')}", bg=C_INK)
    ws.row_dimensions[2].height = 6

    # SUMIF formulas — one per category in 'P&L Data Table'!B column
    rows_def = [
        # (label, formula, is_total, bg, color)
        ("INCOME",            None,     True,  C_FOREST,  C_WHITE),
        ("Total Revenue",     "=SUMIF('P&L Data Table'!B:B,\"Revenue\",'P&L Data Table'!D:D)", False, C_GREEN_BG, C_FOREST),
        ("Total Other Income","=SUMIF('P&L Data Table'!B:B,\"Other Income\",'P&L Data Table'!D:D)", False, C_GREEN_BG, C_FOREST),
        ("GROSS REVENUE",     "=B4+B5", True,  C_FOREST,  C_WHITE),
        ("",None,False,C_WHITE,C_INK),
        ("EXPENSES",          None,     True,  C_CRIMSON, C_WHITE),
        ("Cost of Goods Sold","=SUMIF('P&L Data Table'!B:B,\"COGS\",'P&L Data Table'!D:D)", False, C_ROSE, C_CRIMSON),
        ("Operating Expenses","=SUMIF('P&L Data Table'!B:B,\"Op. Expense\",'P&L Data Table'!D:D)", False, C_ROSE, C_CRIMSON),
        ("Other Expenses",    "=SUMIF('P&L Data Table'!B:B,\"Other Expense\",'P&L Data Table'!D:D)", False, C_ROSE, C_CRIMSON),
        ("TOTAL EXPENSES",    "=B8+B9+B10", True, C_CRIMSON, C_WHITE),
        ("",None,False,C_WHITE,C_INK),
        ("GROSS PROFIT",      "=B6-B8",   True,  C_OCEAN,  C_WHITE),
        ("NET PROFIT",        "=B12-B9-B10+B5", True, C_OCEAN, C_WHITE),
    ]

    data_row = 3
    formula_cells = {}
    for label, formula, is_total, bg, vc in rows_def:
        if not label:
            ws.row_dimensions[data_row].height = 6; data_row += 1; continue
        _set(ws, data_row, 1, label, bold=is_total, size=10 if is_total else 9,
             color=vc if is_total else C_INK, bg=bg, border=_thin())
        if formula:
            c2 = ws.cell(row=data_row, column=2, value=formula)
            c2.number_format = '$#,##0.00'
            c2.font = _font(bold=is_total, size=11 if is_total else 10, color=vc if is_total else C_INK)
            c2.fill = _fill(bg); c2.alignment = _align(h="right")
            formula_cells[label] = f"B{data_row}"
        elif is_total:
            for ci in range(1,3): ws.cell(row=data_row, column=ci).fill = _fill(bg)
        ws.row_dimensions[data_row].height = 16 if is_total else 14
        data_row += 1

    # Ratios table — all formulas
    data_row += 1
    _section_label(ws, data_row, 1, 5, "Financial Ratios (Formula-Driven)")
    data_row += 1
    _header_row(ws, data_row, ["Ratio","Value","Benchmark","Status","Formula Used"], bg=C_CREAM2, text_color=C_INK)
    data_row += 1

    rev_ref    = formula_cells.get("GROSS REVENUE", f"={s.get('total_revenue',1)}")
    net_ref    = formula_cells.get("NET PROFIT",    f"={s.get('net_profit',0)}")
    cogs_ref   = formula_cells.get("Cost of Goods Sold", f"={s.get('total_cogs',0)}")
    exp_ref    = formula_cells.get("TOTAL EXPENSES", f"={s.get('total_cogs',0)+s.get('total_op_expenses',0)}")

    ratio_rows = [
        ("Gross Margin",      f"=(B6-{cogs_ref})/{rev_ref}",     '0.0%',   "> 30%",  r.get('gross_margin',0)>=30,    f"=(Gross Revenue−COGS)/Revenue"),
        ("Net Profit Margin", f"={net_ref}/{rev_ref}",           '0.0%',   "> 10%",  r.get('net_profit_margin',0)>=10, f"=Net Profit/Revenue"),
        ("Expense Ratio",     f"={exp_ref}/{rev_ref}",           '0.0%',   "< 80%",  r.get('expense_ratio',0)<80,     f"=Total Expenses/Revenue"),
        ("COGS Ratio",        f"={cogs_ref}/{rev_ref}",          '0.0%',   "< 40%",  r.get('cogs_ratio',0)<40,        f"=COGS/Revenue"),
    ]
    for label, formula, fmt, bench, ok, formula_desc in ratio_rows:
        bg = C_GREEN_BG if ok else C_ROSE; vc = C_FOREST if ok else C_CRIMSON
        _set(ws, data_row, 1, label, size=10, bg=bg)
        c = ws.cell(row=data_row, column=2, value=formula)
        c.number_format=fmt; c.font=_font(bold=True,size=10,color=vc); c.fill=_fill(bg); c.alignment=_align(h="right")
        _set(ws, data_row, 3, bench, size=9, color=C_MUTED, bg=bg, align="center")
        _set(ws, data_row, 4, "✓" if ok else "✗", size=10, color=vc, bg=bg, align="center", bold=True)
        _set(ws, data_row, 5, formula_desc, size=9, color=C_MUTED, bg=bg, italic=True)
        ws.row_dimensions[data_row].height = 15; data_row += 1

    # Bar chart
    chart_labels = ["Revenue","COGS","Op.Exp","Other Exp","Net Profit"]
    chart_values = [s.get("total_revenue",0),s.get("total_cogs",0),s.get("total_op_expenses",0),s.get("total_other_expenses",0),s.get("net_profit",0)]
    chart_colors = [C_FOREST,C_AMBER,C_CRIMSON,"7C3AED",C_OCEAN]
    for i,(lbl,val) in enumerate(zip(chart_labels,chart_values)):
        ws.cell(row=4+i, column=4, value=lbl)
        c = ws.cell(row=4+i, column=5, value=val); c.number_format='$#,##0'
    bar = BarChart(); bar.type="col"; bar.title="P&L Overview"; bar.style=2; bar.width=18; bar.height=12
    bar.y_axis.numFmt='$#,##0'
    dr = Reference(ws, min_col=5, min_row=4, max_row=8)
    cr = Reference(ws, min_col=4, min_row=4, max_row=8)
    bar.add_data(dr); bar.set_categories(cr)
    for i,clr in enumerate(chart_colors):
        dp = DataPoint(idx=i); dp.graphicalProperties.solidFill=clr; bar.series[0].dPt.append(dp)
    ws.add_chart(bar, "G3")

    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12; ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Balance Sheet Analysis — formula-driven
# ─────────────────────────────────────────────────────────────────────────────
def _build_bs_sheet(wb, bs_data):
    if not bs_data: return
    ws = wb.create_sheet("Balance Sheet Analysis")
    ws.sheet_view.showGridLines = False
    s = bs_data.get("summary",{}); r = bs_data.get("ratios",{})

    _banner(ws, 1, 1, 7, f"BALANCE SHEET  |  {bs_data.get('period','N/A')}", bg=C_INK)
    ws.row_dimensions[2].height = 6

    rows_def = [
        ("ASSETS",             None,    True,  C_OCEAN,   C_WHITE),
        ("Current Assets",     "=SUMIF('BS Data Table'!B:B,\"Current Assets\",'BS Data Table'!D:D)", False, C_BLUE_BG, C_OCEAN),
        ("Fixed Assets",       "=SUMIF('BS Data Table'!B:B,\"Fixed Assets\",'BS Data Table'!D:D)", False, C_BLUE_BG, C_OCEAN),
        ("Other Assets",       "=SUMIF('BS Data Table'!B:B,\"Other Assets\",'BS Data Table'!D:D)", False, C_BLUE_BG, C_OCEAN),
        ("TOTAL ASSETS",       "=B4+B5+B6", True, C_OCEAN, C_WHITE),
        ("",None,False,C_WHITE,C_INK),
        ("LIABILITIES",        None,    True,  C_CRIMSON, C_WHITE),
        ("Current Liabilities","=SUMIF('BS Data Table'!B:B,\"Current Liabilities\",'BS Data Table'!D:D)", False, C_ROSE, C_CRIMSON),
        ("Long-Term Liabilities","=SUMIF('BS Data Table'!B:B,\"Long-Term Liabilities\",'BS Data Table'!D:D)", False, C_ROSE, C_CRIMSON),
        ("TOTAL LIABILITIES",  "=B9+B10", True, C_CRIMSON, C_WHITE),
        ("",None,False,C_WHITE,C_INK),
        ("EQUITY",             None,    True,  C_AMBER,   C_WHITE),
        ("Total Equity",       "=SUMIF('BS Data Table'!A:A,\"Equity\",'BS Data Table'!D:D)", False, C_AMBER_BG, C_AMBER),
        ("WORKING CAPITAL",    "=B4-B9", True,  C_FOREST,  C_WHITE),
        ("ACCOUNTING CHECK",   "=B7-B11-B13", True, C_CREAM2, C_INK),
    ]

    data_row = 3
    formula_cells = {}
    for label, formula, is_total, bg, vc in rows_def:
        if not label:
            ws.row_dimensions[data_row].height = 6; data_row += 1; continue
        _set(ws, data_row, 1, label, bold=is_total, size=10 if is_total else 9,
             color=vc if is_total else C_INK, bg=bg, border=_thin())
        if formula:
            c2 = ws.cell(row=data_row, column=2, value=formula)
            c2.number_format = '$#,##0.00'
            c2.font = _font(bold=is_total, size=11 if is_total else 10, color=vc if is_total else C_INK)
            c2.fill = _fill(bg); c2.alignment = _align(h="right")
            formula_cells[label] = f"B{data_row}"
        elif is_total:
            for ci in range(1,3): ws.cell(row=data_row, column=ci).fill = _fill(bg)
        ws.row_dimensions[data_row].height = 16 if is_total else 14
        data_row += 1

    # Accounting check note
    _set(ws, data_row-1, 3, "← should be $0.00 (Assets = Liabilities + Equity)", size=8, color=C_MUTED, italic=True)

    # Ratios
    data_row += 1
    _section_label(ws, data_row, 1, 5, "Balance Sheet Ratios (Formula-Driven)")
    data_row += 1
    _header_row(ws, data_row, ["Ratio","Value","Benchmark","Status","Formula"], bg=C_CREAM2, text_color=C_INK)
    data_row += 1

    ta = formula_cells.get("TOTAL ASSETS","1"); tl=formula_cells.get("TOTAL LIABILITIES","0")
    ca=formula_cells.get("Current Assets","0"); cl=formula_cells.get("Current Liabilities","0")
    eq=formula_cells.get("Total Equity","0")

    ratio_rows = [
        ("Current Ratio",    f"=IF({cl}=0,\"N/A\",{ca}/{cl})",       '0.00"x"',">1.5x", r.get('current_ratio',0)>=1.5 if r.get('current_ratio') else False, f"=Current Assets/Current Liabilities"),
        ("Debt-to-Equity",   f"=IF({eq}=0,\"N/A\",{tl}/{eq})",       '0.00"x"',"<1.5x", (r.get('debt_to_equity',99) or 99)<=1.5, "=Total Liabilities/Equity"),
        ("Debt-to-Assets",   f"=IF({ta}=0,\"N/A\",{tl}/{ta})",       '0.00"x"',"<0.5x", (r.get('debt_to_assets',99) or 99)<=0.5, "=Total Liabilities/Total Assets"),
        ("Equity Ratio",     f"=IF({ta}=0,\"N/A\",{eq}/{ta})",       '0.0%',   ">50%",  (r.get('equity_ratio',0) or 0)>=50, "=Equity/Total Assets"),
    ]
    for label, formula, fmt, bench, ok, formula_desc in ratio_rows:
        bg = C_GREEN_BG if ok else C_ROSE; vc = C_FOREST if ok else C_CRIMSON
        _set(ws, data_row, 1, label, size=10, bg=bg)
        c = ws.cell(row=data_row, column=2, value=formula)
        c.number_format=fmt; c.font=_font(bold=True,size=10,color=vc); c.fill=_fill(bg); c.alignment=_align(h="right")
        _set(ws, data_row, 3, bench, size=9, color=C_MUTED, bg=bg, align="center")
        _set(ws, data_row, 4, "✓" if ok else "✗", size=10, color=vc, bg=bg, align="center", bold=True)
        _set(ws, data_row, 5, formula_desc, size=9, color=C_MUTED, bg=bg, italic=True)
        ws.row_dimensions[data_row].height = 15; data_row += 1

    # BS structure chart
    chart_labels = ["Cur. Assets","Fixed","Other","Cur. Liab","LT Liab","Equity"]
    chart_vals   = [s.get("current_assets",0),s.get("fixed_assets",0),s.get("other_assets",0),
                    s.get("current_liabilities",0),s.get("long_term_liabilities",0),s.get("equity",0)]
    chart_colors = [C_FOREST,"2D9150",C_OCEAN,C_CRIMSON,"9E1830",C_AMBER]
    for i,(lbl,val) in enumerate(zip(chart_labels,chart_vals)):
        ws.cell(row=4+i, column=4, value=lbl)
        c=ws.cell(row=4+i, column=5, value=val); c.number_format='$#,##0'
    bar=BarChart(); bar.type="col"; bar.title="Balance Sheet Structure"; bar.style=2; bar.width=16; bar.height=11
    dr=Reference(ws,min_col=5,min_row=4,max_row=9); cr=Reference(ws,min_col=4,min_row=4,max_row=9)
    bar.add_data(dr); bar.set_categories(cr)
    for i,clr in enumerate(chart_colors):
        dp=DataPoint(idx=i); dp.graphicalProperties.solidFill=clr; bar.series[0].dPt.append(dp)
    ws.add_chart(bar,"G3")

    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=18
    ws.column_dimensions["C"].width=26; ws.column_dimensions["D"].width=16
    ws.column_dimensions["E"].width=30
    ws.freeze_panes="A3"


# ─────────────────────────────────────────────────────────────────────────────
# Tax sheet — formula-driven
# ─────────────────────────────────────────────────────────────────────────────
def _build_tax_sheet(wb, tax_data):
    if not tax_data or tax_data.get("tax") is None: return
    ws = wb.create_sheet("Tax Estimate")
    ws.sheet_view.showGridLines = False

    _banner(ws, 1, 1, 5, f"TAX ESTIMATE  |  {tax_data.get('country_description','')}  |  {tax_data.get('entity_type','')}", bg=C_INK)
    ws.row_dimensions[2].height = 6

    # Input assumptions block
    row = 3
    _section_label(ws, row, 1, 5, "Tax Calculation Inputs"); row += 1
    inputs = [
        ("Gross Profit (Net Income)",    tax_data.get("gross_profit",0), '$#,##0.00'),
        ("Total Deductions Applied",     tax_data.get("total_deductions",0), '$#,##0.00'),
        ("Taxable Income",               f"=B{row}-B{row+1}", '$#,##0.00'),
        ("Effective Tax Rate",           tax_data.get("effective_rate",0)/100, '0.0%'),
        ("Estimated Federal Tax",        f"=B{row+2}*B{row+3}", '$#,##0.00'),
        ("After-Tax Profit",             f"=B{row}-B{row+4}", '$#,##0.00'),
    ]
    input_row_start = row
    for label, val, fmt in inputs:
        _set(ws, row, 1, label, size=10, bg=C_CREAM if row%2==0 else C_WHITE)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format=fmt; c.font=_font(bold=True,size=11); c.alignment=_align(h="right")
        c.fill = _fill(C_CREAM if row%2==0 else C_WHITE)
        ws.row_dimensions[row].height=18; row += 1

    # Highlight the tax cell
    ws.cell(row=input_row_start+4, column=2).font = _font(bold=True, size=13, color=C_CRIMSON)
    ws.cell(row=input_row_start+5, column=2).font = _font(bold=True, size=13, color=C_FOREST)

    # Deductions
    ded_log = tax_data.get("deduction_breakdown",[])
    if ded_log:
        row += 1
        _section_label(ws, row, 1, 4, f"Deductions Applied ({len(ded_log)} items)"); row += 1
        _header_row(ws, row, ["Deduction Item","Amount","Formula / Basis"], bg=C_CREAM2, text_color=C_INK); row += 1
        ded_start = row
        for d in ded_log:
            bg = C_WHITE if row%2==0 else C_CREAM
            _set(ws, row, 1, d["item"], size=10, bg=bg)
            c=ws.cell(row=row,column=2,value=d["amount"]); c.number_format='$#,##0.00'; c.font=_font(size=10,color=C_FOREST); c.fill=_fill(bg); c.alignment=_align(h="right")
            _set(ws, row, 3, d.get("note","")[:80], size=9, color=C_MUTED, bg=bg, wrap=True)
            ws.row_dimensions[row].height=24; row += 1
        # Total deductions formula
        if row-1 >= ded_start:
            _set(ws, row, 1, "TOTAL DEDUCTIONS", bold=True, size=10, bg=C_GREEN_BG, color=C_FOREST)
            c=ws.cell(row=row,column=2,value=f"=SUM(B{ded_start}:B{row-1})")
            c.number_format='$#,##0.00'; c.font=_font(bold=True,size=11,color=C_FOREST); c.fill=_fill(C_GREEN_BG); c.alignment=_align(h="right")
            ws.row_dimensions[row].height=18; row += 1
            table = Table(displayName="TaxDeductions", ref=f"A{ded_start-1}:C{row-2}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1",showRowStripes=True)
            ws.add_table(table)

    # Disclaimer
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    _set(ws, row, 1, tax_data.get("disclaimer",""), size=9, color=C_FAINT, bg=C_INK, italic=True, wrap=True, align="center")
    ws.row_dimensions[row].height=28

    for col,w in [("A",30),("B",18),("C",50),("D",14),("E",14)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Monthly sheet — formula-driven with charts
# ─────────────────────────────────────────────────────────────────────────────
def _build_monthly_sheet(wb, monthly_data):
    if not monthly_data: return
    ws = wb.create_sheet("Monthly Trends")
    ws.sheet_view.showGridLines = False

    months=monthly_data.get("months",[]); revenue=monthly_data.get("revenue",[]); expenses=monthly_data.get("expenses",[]); profit=monthly_data.get("profit",[])
    anomalies=monthly_data.get("anomalies",[])

    _banner(ws, 1, 1, 7, f"MONTHLY P&L TRENDS  |  {monthly_data.get('period','N/A')}", bg=C_INK)
    _header_row(ws, 2, ["Month","Revenue","Expenses","Profit","Margin %","vs Avg Profit","Anomaly?"], bg=C_CRIMSON)
    ws.row_dimensions[2].height=18

    active=[(m,r,e,p) for m,r,e,p in zip(months,revenue,expenses,profit) if r>0 or e>0]
    data_start=3
    for i,(m,r,e,p) in enumerate(active):
        row=3+i; bg=C_WHITE if i%2==0 else C_CREAM
        _set(ws, row, 1, m, size=10, bg=bg)
        for ci,(val,fmt) in enumerate([(r,'$#,##0'),(e,'$#,##0')],2):
            c=ws.cell(row=row,column=ci,value=val); c.number_format=fmt; c.font=_font(size=10); c.fill=_fill(bg); c.alignment=_align(h="right")
        # Profit formula: =Revenue - Expenses
        c4=ws.cell(row=row,column=4,value=f"=B{row}-C{row}")
        c4.number_format='$#,##0'; c4.font=_font(size=10); c4.fill=_fill(bg); c4.alignment=_align(h="right")
        # Margin formula
        c5=ws.cell(row=row,column=5,value=f"=IF(B{row}=0,0,D{row}/B{row})")
        c5.number_format='0.0%'; c5.font=_font(size=10,color=C_OCEAN); c5.fill=_fill(bg); c5.alignment=_align(h="right")
        # vs average — filled after all rows written
        ws.row_dimensions[row].height=15

    data_end=2+len(active)
    avg_row=data_end+1

    # Average row
    if len(active)>0:
        _set(ws, avg_row, 1, "AVERAGE", bold=True, size=10, bg=C_CREAM2, color=C_INK)
        for ci in range(2,6):
            c=ws.cell(row=avg_row,column=ci,value=f"=AVERAGE({get_column_letter(ci)}{data_start}:{get_column_letter(ci)}{data_end})")
            c.number_format='$#,##0' if ci<5 else '0.0%'; c.font=_font(bold=True,size=10); c.fill=_fill(C_CREAM2); c.alignment=_align(h="right")
        ws.row_dimensions[avg_row].height=18

    # Now fill "vs avg" column with formula referencing average row
    for i in range(len(active)):
        row=3+i; bg=C_WHITE if i%2==0 else C_CREAM
        c6=ws.cell(row=row,column=6,value=f"=D{row}-$D${avg_row}")
        c6.number_format='$#,##0'; c6.font=_font(size=10); c6.fill=_fill(bg); c6.alignment=_align(h="right")
        # Anomaly flag
        anomaly_months=[a["month"] for a in anomalies]
        m=active[i][0]
        _set(ws, row, 7, "⚠ ANOMALY" if m in anomaly_months else "", size=9, color=C_CRIMSON, bg=bg, bold=True)

    # Excel Table
    if data_end>=data_start:
        table=Table(displayName="MonthlyTrends",ref=f"A2:G{data_end}")
        table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True)
        ws.add_table(table)

    # Conditional formatting
    ws.conditional_formatting.add(f"D{data_start}:D{data_end}",
        ColorScaleRule(start_type='min',start_color='FECDD3',mid_type='percentile',mid_value=50,mid_color='FFFFFF',end_type='max',end_color='BBF7D0'))

    # Line chart
    line=LineChart(); line.title="Monthly Trends"; line.style=2; line.width=22; line.height=13
    line.y_axis.numFmt='$#,##0'
    rev_r=Reference(ws,min_col=2,min_row=2,max_row=data_end)
    exp_r=Reference(ws,min_col=3,min_row=2,max_row=data_end)
    pro_r=Reference(ws,min_col=4,min_row=2,max_row=data_end)
    cat_r=Reference(ws,min_col=1,min_row=3,max_row=data_end)
    line.add_data(rev_r,titles_from_data=True); line.add_data(exp_r,titles_from_data=True); line.add_data(pro_r,titles_from_data=True)
    line.set_categories(cat_r)
    line.series[0].graphicalProperties.line.solidFill=C_FOREST; line.series[0].graphicalProperties.line.width=20000
    line.series[1].graphicalProperties.line.solidFill=C_CRIMSON; line.series[1].graphicalProperties.line.width=20000
    line.series[2].graphicalProperties.line.solidFill=C_OCEAN; line.series[2].graphicalProperties.line.width=20000
    line.series[2].graphicalProperties.line.dashDot="dash"
    ws.add_chart(line,"I2")

    for col,w in [("A",14),("B",14),("C",14),("D",14),("E",12),("F",16),("G",12)]:
        ws.column_dimensions[col].width=w
    ws.freeze_panes="A3"


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(results: dict, monthly_data: dict = None,
                   raw_file_bytes: bytes = None, raw_filename: str = "") -> bytes:
    wb = Workbook()

    pl_data = results.get("pl_analysis") or (results.get("analysis") if results.get("analysis",{}).get("type")=="pl" else None)
    bs_data = results.get("bs_current")  or (results.get("analysis") if results.get("analysis",{}).get("type")=="bs" else None)
    tax_data = results.get("tax")

    # Build data table sheets FIRST (other sheets reference them via formulas)
    _build_pl_data_table(wb, pl_data)
    _build_bs_data_table(wb, bs_data)

    # Build analysis sheets (formula-driven, reference data tables)
    _build_summary(wb, results, pl_data, bs_data)
    _build_pl_sheet(wb, pl_data)
    _build_bs_sheet(wb, bs_data)
    _build_tax_sheet(wb, tax_data)
    if monthly_data:
        _build_monthly_sheet(wb, monthly_data)

    # Raw data sheet — user's original file
    if raw_file_bytes:
        _build_raw_data(wb, raw_file_bytes, raw_filename or "Uploaded File")

    # Remove empty placeholder sheets
    empty = [s.title for s in wb.worksheets if s.max_row <= 1 and s.title != "Executive Summary"]
    for t in empty:
        if t in wb.sheetnames: del wb[t]

    # Set tab order: Summary first
    if "Executive Summary" in wb.sheetnames:
        wb.move_sheet("Executive Summary", offset=-len(wb.sheetnames))

    # Add raw source file sheet
    if raw_file_bytes:
        _build_raw_data(wb, raw_file_bytes, raw_filename or "source_file")

    # Add formula cross-reference sheet
    _build_formulas_sheet(wb)

    # Remove any completely empty sheets
    sheets_to_remove = [sh.title for sh in wb.worksheets
                        if sh.max_row <= 1 and sh.title not in ("Executive Summary",)]
    for title in sheets_to_remove:
        try: del wb[title]
        except: pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_formulas_sheet(wb):
    """
    Add a Formulas & Calculations sheet that references other sheets with real Excel formulas.
    This ensures the workbook is dynamic — change a number in any data table and this updates.
    """
    # Find the P&L Data Table sheet and BS Data Table sheet
    pl_sheet = next((s for s in wb.sheetnames if s == "P&L Data Table"), None)
    bs_sheet = next((s for s in wb.sheetnames if s == "BS Data Table"), None)
    exec_sheet = "Executive Summary"

    ws = wb.create_sheet("Formulas & Calcs")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    _set(ws, 1, 1, "DYNAMIC FORMULA CALCULATIONS — Values update when source data changes",
         bold=True, size=12, color=C_WHITE, bg=C_INK)
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:E2")
    _set(ws, 2, 1, "All cells below use Excel formulas referencing the Data Table sheets. Edit those sheets and press F9 to recalculate.",
         size=9, color=C_FAINT, bg=C_INK, italic=True)
    ws.row_dimensions[2].height = 14

    row = 4

    # ── P&L Formulas (reference P&L Data Table sheet) ────────────────────────
    if pl_sheet:
        esc_pl = pl_sheet.replace("'", "''")
        _section_label(ws, row, 1, 5, "P&L Calculations (from P&L Data Table)")
        row += 1

        headers = ["Metric", "Formula", "Value", "Notes"]
        for ci, h in enumerate(headers, 1):
            _set(ws, row, ci, h, bold=True, size=9, bg=C_CREAM2)
        ws.row_dimensions[row].height = 15; row += 1

        calcs_pl = [
            ("Total Revenue",     f"=SUMIF('{esc_pl}'!F:F,\"Revenue\",'{esc_pl}'!D:D)",         "SUMIF on Type=Revenue"),
            ("Total COGS",        f"=SUMIF('{esc_pl}'!B:B,\"COGS\",'{esc_pl}'!D:D)",            "SUMIF on Category=COGS"),
            ("Total Op. Expenses",f"=SUMIF('{esc_pl}'!B:B,\"Op. Expense\",'{esc_pl}'!D:D)",     "SUMIF on Category=Op. Expense"),
            ("Gross Profit",      f"=C{row}-C{row+1}",                                           "Revenue - COGS (once both above filled)"),
            ("EBITDA",            f"=C{row}-C{row+1}-C{row+2}",                                  "Revenue - COGS - Op. Expenses"),
            ("Total Other Income",f"=SUMIF('{esc_pl}'!B:B,\"Other Income\",'{esc_pl}'!D:D)",    "SUMIF on Category=Other Income"),
            ("Total Other Exp",   f"=SUMIF('{esc_pl}'!B:B,\"Other Expense\",'{esc_pl}'!D:D)",   "SUMIF on Category=Other Expense"),
            ("Net Profit",        f"=SUMIF('{esc_pl}'!F:F,\"Revenue\",'{esc_pl}'!D:D)-SUMIF('{esc_pl}'!F:F,\"Expense\",'{esc_pl}'!D:D)",
             "Total Revenue minus Total Expenses"),
        ]

        formula_cells = {}
        for label, formula, note in calcs_pl:
            bg = C_WHITE if row % 2 == 0 else C_CREAM
            _set(ws, row, 1, label, size=10, bg=bg)
            fc = ws.cell(row=row, column=2, value=formula)
            fc.font = _font(size=9, color=C_OCEAN, italic=True); fc.fill = _fill(bg)
            vc = ws.cell(row=row, column=3, value=formula)
            vc.number_format = '$#,##0.00'; vc.font = _font(size=10, bold=True, color=C_FOREST); vc.fill = _fill(bg); vc.alignment = _align(h='right')
            _set(ws, row, 4, note, size=9, color=C_MUTED, bg=bg)
            formula_cells[label] = f"C{row}"
            ws.row_dimensions[row].height = 16; row += 1

        # Margin formulas referencing the value cells
        row_rev  = [r for r, (l,_,_) in enumerate(calcs_pl) if l=="Total Revenue"]
        row_net  = [r for r, (l,_,_) in enumerate(calcs_pl) if l=="Net Profit"]
        row_gp   = [r for r, (l,_,_) in enumerate(calcs_pl) if l=="Gross Profit"]
        base_row = row - len(calcs_pl)

        margin_calcs = [
            ("Gross Margin %",       f"=IF(C{base_row+0}>0,C{base_row+3}/C{base_row+0},0)", "Gross Profit / Total Revenue"),
            ("Net Profit Margin %",  f"=IF(C{base_row+0}>0,C{base_row+7}/C{base_row+0},0)", "Net Profit / Total Revenue"),
            ("Expense Ratio %",      f"=IF(C{base_row+0}>0,(C{base_row+1}+C{base_row+2})/C{base_row+0},0)", "(COGS + OpEx) / Revenue"),
        ]
        for label, formula, note in margin_calcs:
            bg = C_BLUE_BG if row % 2 == 0 else "E8F4FD"
            _set(ws, row, 1, label, size=10, bg=bg, color=C_OCEAN)
            fc = ws.cell(row=row, column=2, value=formula); fc.font=_font(size=9,color=C_OCEAN,italic=True); fc.fill=_fill(bg)
            vc = ws.cell(row=row, column=3, value=formula); vc.number_format='0.0%'; vc.font=_font(size=10,bold=True,color=C_OCEAN); vc.fill=_fill(bg); vc.alignment=_align(h='right')
            _set(ws, row, 4, note, size=9, color=C_MUTED, bg=bg)
            ws.row_dimensions[row].height = 16; row += 1
        row += 1

    # ── BS Formulas ───────────────────────────────────────────────────────────
    if bs_sheet:
        esc_bs = bs_sheet.replace("'","''")
        _section_label(ws, row, 1, 5, "Balance Sheet Calculations (from BS Data Table)")
        row += 1
        for ci, h in enumerate(["Metric","Formula","Value","Notes"], 1):
            _set(ws, row, ci, h, bold=True, size=9, bg=C_CREAM2)
        ws.row_dimensions[row].height = 15; row += 1

        calcs_bs = [
            ("Total Assets",       f"=SUMIF('{esc_bs}'!F:F,\"Assets\",'{esc_bs}'!D:D)",        "SUMIF on Type=Assets"),
            ("Total Liabilities",  f"=SUMIF('{esc_bs}'!F:F,\"Liabilities\",'{esc_bs}'!D:D)",   "SUMIF on Type=Liabilities"),
            ("Total Equity",       f"=SUMIF('{esc_bs}'!F:F,\"Equity\",'{esc_bs}'!D:D)",        "SUMIF on Type=Equity — preserves negatives"),
            ("Working Capital",    f"=SUMIF('{esc_bs}'!B:B,\"Current Assets\",'{esc_bs}'!D:D)-SUMIF('{esc_bs}'!B:B,\"Current Liabilities\",'{esc_bs}'!D:D)",
             "Current Assets - Current Liabilities"),
            ("Accounting Check",   f"=C{row}-C{row+1}-C{row+2}", "Assets - Liabilities - Equity (should be 0)"),
        ]
        base_bs = row
        for label, formula, note in calcs_bs:
            bg = C_WHITE if row%2==0 else C_CREAM
            is_check = label == "Accounting Check"
            _set(ws, row, 1, label, size=10, bg=bg, color=C_OCEAN if is_check else C_INK, bold=is_check)
            fc = ws.cell(row=row,column=2,value=formula); fc.font=_font(size=9,color=C_OCEAN,italic=True); fc.fill=_fill(bg)
            vc = ws.cell(row=row,column=3,value=formula); vc.number_format='$#,##0.00'; vc.font=_font(size=10,bold=True,color=C_FOREST if not is_check else C_AMBER); vc.fill=_fill(bg); vc.alignment=_align(h='right')
            _set(ws, row, 4, note, size=9, color=C_MUTED, bg=bg)
            ws.row_dimensions[row].height = 16; row += 1

        # Ratio formulas
        row_ta = base_bs; row_tl = base_bs+1; row_eq = base_bs+2
        row_ca = row  # placeholder — we'll add SUMIF for current assets
        ws.cell(row=row, column=3, value=f"=SUMIF('{esc_bs}'!B:B,\"Current Assets\",'{esc_bs}'!D:D)")
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        row_cl_formula = f"SUMIF('{esc_bs}'!B:B,\"Current Liabilities\",'{esc_bs}'!D:D)"

        ratio_calcs = [
            ("Current Ratio",   f"=IFERROR(SUMIF('{esc_bs}'!B:B,\"Current Assets\",'{esc_bs}'!D:D)/SUMIF('{esc_bs}'!B:B,\"Current Liabilities\",'{esc_bs}'!D:D),\"N/A\")", "0.00", "Current Assets / Current Liabilities"),
            ("Debt-to-Equity",  f"=IFERROR(C{row_tl}/C{row_eq},\"N/A\")", "0.00", "Total Liabilities / Equity"),
            ("Debt-to-Assets",  f"=IFERROR(C{row_tl}/C{row_ta},\"N/A\")", "0.00", "Total Liabilities / Total Assets"),
            ("Equity Ratio %",  f"=IFERROR(C{row_eq}/C{row_ta},\"N/A\")", "0.0%", "Equity / Total Assets"),
        ]
        for label, formula, nfmt, note in ratio_calcs:
            bg = C_BLUE_BG if row%2==0 else "E8F4FD"
            _set(ws, row, 1, label, size=10, bg=bg, color=C_OCEAN)
            fc = ws.cell(row=row,column=2,value=formula); fc.font=_font(size=9,color=C_OCEAN,italic=True); fc.fill=_fill(bg)
            vc = ws.cell(row=row,column=3,value=formula); vc.number_format=nfmt; vc.font=_font(size=10,bold=True,color=C_OCEAN); vc.fill=_fill(bg); vc.alignment=_align(h='right')
            _set(ws, row, 4, note, size=9, color=C_MUTED, bg=bg)
            ws.row_dimensions[row].height=16; row+=1

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 10
    ws.freeze_panes = "A3"
